import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import zipfile 
from utils import log_error

# --- UPDATED COLUMNS ---
DEFAULT_COLS = ["S.No.", "Name", "Email", "Phone", "Experience", "Status"]

def _apply_dropdown_validation(ws):
    """Applies data validation dropdowns to the 'Status' column."""
    try:
        # Status options including new Re-Applicant statuses
        STATUS_OPTIONS = [
            "New Applicant", "Re-Applicant (Updated)", "Duplicate (Ignored)",
            "Accepted", "Rejected", "On Hold", 
            "Interview Scheduled", "Pending Review"
        ]
        
        dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
        dv.error = "Your entry is not in the list."
        dv.errorTitle = "Invalid Entry"
        ws.add_data_validation(dv)
        
        status_col_index = -1
        for i, cell in enumerate(ws[1]): 
            if cell.value == "Status":
                status_col_index = i + 1
                break
        
        if status_col_index != -1:
            status_col_letter = get_column_letter(status_col_index)
            dv.add(f'{status_col_letter}2:{status_col_letter}1048576')
        
    except Exception as e:
        log_error(f"Validation error: {e}")

def validate_or_create_excel(path: str):
    """Ensures the Excel file exists and has correct headers."""
    if not os.path.exists(path):
        _create_new_excel(path)
        return

    try:
        wb = load_workbook(path)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        
        if existing_headers == DEFAULT_COLS:
            _apply_dropdown_validation(ws)
            wb.save(path)
            return
        else:
            # Simple check: if Experience is missing, we might need to recreate or append
            # For robustness, we back up and recreate if headers don't match exactly
            log_error(f"Header mismatch. Backing up {path} and recreating.")
            os.rename(path, path + ".bak_ver")
            _create_new_excel(path)
            
    except Exception:
        if os.path.exists(path): os.remove(path)
        _create_new_excel(path)

def _create_new_excel(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"
    ws.append(DEFAULT_COLS)
    _apply_dropdown_validation(ws)
    wb.save(path)

def get_next_serial_number(path: str) -> int:
    if not os.path.exists(path): return 1
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in range(ws.max_row, 1, -1):
            val = ws.cell(row=row, column=1).value
            if val is not None: return int(val) + 1
        return 1
    except Exception:
        return 1

def append_row(path: str, data: dict, status: str = "New Applicant") -> int:
    """
    Append a data row to Excel. Returns the new serial number.
    data keys: Name, Email, Phone, Experience
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        serial_num = get_next_serial_number(path)
        
        email_val = data.get("Email", "")
        if isinstance(email_val, list): email_val = ", ".join(email_val)
        
        row = [
            serial_num,
            data.get("Name", ""),
            email_val,
            data.get("Phone", ""),
            data.get("Experience", "0"), # New Column
            status
        ]
        ws.append(row)
        wb.save(path)
        return serial_num
    except Exception as e:
        log_error(f"Excel append failed: {e}")
        raise

def read_all_rows(path: str):
    """Return list of value-tuples from Excel (skips header)."""
    if not os.path.exists(path): return []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row_data = list(r)
            # Pad or truncate to match DEFAULT_COLS length
            if len(row_data) < len(DEFAULT_COLS):
                row_data.extend([""] * (len(DEFAULT_COLS) - len(row_data)))
            rows.append(tuple(row_data[:len(DEFAULT_COLS)]))
        return rows
    except Exception:
        return []

def update_status(path: str, serial_num: int, new_status: str) -> bool:
    if not os.path.exists(path): return False
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        status_col = -1
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Status":
                status_col = idx
                break
        
        if status_col == -1: return False

        for row in ws.iter_rows(min_row=2):
            if row[0].value == serial_num:
                row[status_col-1].value = new_status
                wb.save(path)
                return True
        return False
    except Exception as e:
        log_error(f"Status update failed: {e}")
        return False

def export_by_status(excel_path: str, destination_folder: str):
    """Exports rows to separate files based on Status."""
    if not os.path.exists(excel_path): return []

    wb_main = load_workbook(excel_path, data_only=True)
    ws_main = wb_main.active
    headers = [c.value for c in ws_main[1]]
    
    try:
        status_idx = headers.index("Status")
    except ValueError:
        return []

    status_map = {} # status -> list of rows
    
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        status = row[status_idx]
        if status:
            if status not in status_map: status_map[status] = []
            status_map[status].append(row)
            
    created_files = []
    for status, rows in status_map.items():
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.append(headers)
        for r in rows: ws_new.append(r)
        
        safe_name = "".join([c if c.isalnum() else "_" for c in status])
        fname = os.path.join(destination_folder, f"{safe_name}_Candidates.xlsx")
        wb_new.save(fname)
        created_files.append(fname)
        
    return created_files
